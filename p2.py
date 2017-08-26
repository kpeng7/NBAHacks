from collections import defaultdict
import random
from openpyxl import load_workbook
import copy

#dictionary of date : [(winning team, losing team, score)]
GAME_DATA = defaultdict(list)
WESTERN_CONF = None
EASTERN_CONF = None

class Team():
    name = ""
    conference_rank = 0     #set in conference object
    division_rank = 0       #set in division object
    games_won = float(0)
    games_played = float(0)
    division = ""
    conference = ""
    division_games_won = float(0)
    division_games_played = float(0)
    conference_games_won = float(0)
    conference_games_played = float(0)
    opponents = {}          #dict of opponents names (str): [number of wins, number of losses]
    points_scored = float(0)
    points_allowed = float(0)
    eliminated = "Playoffs"

    #constructor for Team object
    def __init__(self, name, division, conference, opponents_list, opponents = None):
        if opponents == None:
            opponents = {}
        self.opponents = opponents
        self.name = name
        self.division = division
        self.conference = conference
        for opponent in opponents_list:
            self.opponents[opponent] = [0, 0] #FORMAT = [W, L]

    #adds a game played to the team
    def addGame(self, opponent, outcome):
        won = False
        if outcome[0] > outcome[1]:
            won = True
        self.points_scored += outcome[0]
        self.points_allowed += outcome[1]
        if won:
            self.games_won += 1
            self.opponents[opponent.name][0] += 1
        else:
            self.opponents[opponent.name][1] += 1
        if self.division == opponent.division:
            self.division_games_played += 1
            if won:
                self.division_games_won += 1
        if self.conference == opponent.conference:
            self.conference_games_played += 1
            if won:
                self.conference_games_won += 1
        self.games_played += 1

    def winRest(self):
        new_team = copy.deepcopy(self)
        new_team.games_won = new_team.games_won + (82.0 - new_team.games_played)
        new_team.games_played = 82.0
        return new_team

    def loseRest(self):
        new_team = copy.deepcopy(self)
        new_team.games_played = 82.0
        return new_team

class Group():
    name = ""
    teams = []
    team_names = []
    group_leader = None
    rankings = {} #dict of rank : team

    #constructor for Group object
    def __init__(self, name, teams, team_names = None, rankings = None):
        if team_names == None:
            team_names = []
        if rankings == None:
            rankings = {}
        self.team_names = team_names
        self.rankings = rankings
        self.name = name
        self.teams = teams
        for team in teams:
            self.team_names.append(team.name)

    def rankTeams(self, conf, opp_conf):
        ranked_list = self.settleTie(self.teams, conf, opp_conf)
        check = {}
        for i in range(len(ranked_list)):
            self.rankings[i + 1] = ranked_list[i]
            check[i + 1] = ranked_list[i].name
            if not self.name == "West" and not self.name == "East":
                ranked_list[i].division_rank = i + 1
            else:
                ranked_list[i].conference_rank = i + 1

    def rankByOverallWinPercentage(self, teams):
        out = []
        ranked_dict = defaultdict(list)
        for team in teams:
            ranked_dict[float(team.games_won)/float(team.games_played)].append(team)
        sorted_keys = sorted(ranked_dict.keys())
        for key in reversed(sorted_keys):
            out.append(ranked_dict[key])
        return out

    def divisionLeader(self, teams):
        division_leaders = []
        non_division_leaders = []
        for team in teams:
            if team.division_rank == 1:
                division_leaders.append(team)
            else:
                non_division_leaders.append(team)
        if not division_leaders:
            return [non_division_leaders]
        elif not non_division_leaders:
            return [division_leaders]
        else:
            out = []
            out.append(division_leaders)
            out.append(non_division_leaders)
            return out

    def rankByRecordAgainstAll(self, teams):
        out = []
        records = defaultdict(int)
        total_games = defaultdict(int)
        record_to_team = defaultdict(list)
        for i in range(len(teams) - 1):
            team1 = teams[i]
            for j in range(i + 1, len(teams)):
                team2 = teams[j]
                (wins, loss) = team1.opponents[team2.name]
                total_games[team1] += wins + loss
                total_games[team2] += wins + loss
                records[team1] += wins
                records[team2] += loss
            if total_games[team1] == 0:
                return [teams]
        for team in teams:
            record_to_team[records[team] / total_games[team]].append(team1)
        sorted_keys = sorted(record_to_team.keys())
        for key in reversed(sorted_keys):
            out.append(record_to_team[key])
        return out

    def rankByDivWonLostPercentage(self, teams):
        out = []
        div_won_lost = defaultdict(list)
        division = teams[0].division
        for team in teams:
            if not team.division == division:
                return [teams]
            div_won_lost[team.division_games_won / team.division_games_played].append(team)
        sorted_keys = sorted(div_won_lost.keys())
        for key in reversed(sorted_keys):
            out.append(div_won_lost[key])
        return out

    def rankByConfWonLostPercentage(self, teams):
        out = []
        conf_won_lost = defaultdict(list)
        for team in teams:
            conf_won_lost[team.conference_games_won / team.conference_games_played].append(team)
        sorted_keys = sorted(conf_won_lost.keys())
        for key in reversed(sorted_keys):
            out.append(conf_won_lost[key])
        return out

    def rankByPlayoffEligibleInConf(self, teams, conf):
        out = []
        eligible_teams = []
        for team in conf.teams:
            if team.eliminated == "Playoffs":
                eligible_teams.append(team)
        record_to_team = defaultdict(list)
        for team in teams:
            total_wins = 0.0
            total_losses = 0.0
            for eligible in eligible_teams:
                if not eligible.name == team.name:
                    total_wins += team.opponents[eligible.name][0]
                    total_losses += team.opponents[eligible.name][1]
            record_to_team[total_wins/(total_wins + total_losses)].append(team)
        sorted_keys = sorted(record_to_team.keys())
        for key in reversed(sorted_keys):
            out.append(record_to_team[key])
        return out

    def rankByPointDifferential(self, teams):
        out = []
        point_dif = defaultdict(list)
        for team in teams:
            point_dif[team.points_scored - team.points_allowed].append(team)
        sorted_keys = sorted(point_dif.keys())
        for key in reversed(sorted_keys):
            out.append(point_dif[key])
        return out

    def randomize(self, teams):
        random.shuffle(teams)
        return [[team] for team in teams]

    def settleTie(self, tied_teams, conf, opp_conf):
        if len(tied_teams) == 1:
            return tied_teams
        elif len(tied_teams) == 2:
            list_of_checks = [self.rankByOverallWinPercentage(tied_teams), \
                              self.rankByRecordAgainstAll(tied_teams), \
                              self.divisionLeader(tied_teams), \
                              self.rankByDivWonLostPercentage(tied_teams), \
                              self.rankByConfWonLostPercentage(tied_teams), \
                              self.rankByPlayoffEligibleInConf(tied_teams, conf), \
                              self.rankByPlayoffEligibleInConf(tied_teams, opp_conf), \
                              self.rankByPointDifferential(tied_teams), \
                              self.randomize(tied_teams)]
            for teams_list in list_of_checks:
                if not len(teams_list) == 1:
                    out = []
                    for teams in teams_list:
                        out.extend(self.settleTie(teams, conf, opp_conf))
                    return out
        else:
            list_of_checks = [self.rankByOverallWinPercentage(tied_teams), \
                              self.divisionLeader(tied_teams), \
                              self.rankByRecordAgainstAll(tied_teams), \
                              self.rankByDivWonLostPercentage(tied_teams), \
                              self.rankByConfWonLostPercentage(tied_teams), \
                              self.rankByPlayoffEligibleInConf(tied_teams, conf), \
                              self.rankByPointDifferential(tied_teams), \
                              self.randomize(tied_teams)]
            for teams_list in list_of_checks:
                if not len(teams_list) == 1:
                    out = []
                    for teams in teams_list:
                        try:
                            out.extend(self.settleTie(teams, conf, opp_conf))
                        except:
                            out.extend(self.settleTie(teams, conf, opp_conf))
                    return out

class Division(Group):

    #constructor for division object
    def __init__(self, name, teams):
        Group.__init__(self, name, teams)

class Conference(Group):
    divisions = []
    #constructor for conference object
    def __init__(self, name, teams, divisions):
        Group.__init__(self, name, teams)
        self.divisions = divisions

    def rankTeams(self):
        for division in self.divisions:
            if self.name == "West":
                division.rankTeams(WESTERN_CONF, EASTERN_CONF)
            else:
                division.rankTeams(EASTERN_CONF, WESTERN_CONF)
        if self.name == "West":
            Group.rankTeams(self, WESTERN_CONF, EASTERN_CONF)
        else:
            Group.rankTeams(self, EASTERN_CONF, WESTERN_CONF)

def generateTeams(ws):
    team_list = {}
    opponents = []
    #returns dictionary of team name : team object
    for row in ws.iter_rows(range_string="A2:C31"):
        opponents.append(row[0].value)
    for row in ws.iter_rows(range_string="A2:C31"):
        team1 = Team(row[0].value, row[1].value, row[2].value, opponents)
        team_list[row[0].value] = team1
    return team_list

def generateDivisions(team_list):
    #returns dictionary of division name : division object
    division_list = {}
    for team_name in team_list:
        team = team_list[team_name]
        current_division = team.division
        if not division_list.has_key(current_division):
            teams_division_list = []
            for team_name_temp in team_list:
                team_temp = team_list[team_name_temp]
                if team_temp.division == current_division:
                    teams_division_list.append(team_temp)
            division_temp = Division(current_division, teams_division_list)
            division_list[current_division] = division_temp
    return division_list

def generateConference(team_list, divisions_list):
    #returns dictionary of conference name : conference object
    conference_list = {}
    teams_west_list = []
    teams_east_list = []
    divisions_west_list = []
    divisions_east_list = []
    for team_name in team_list:
        team = team_list[team_name]
        if team.conference == "West":
            teams_west_list.append(team)
            if not divisions_list[team.division] in divisions_west_list:
                divisions_west_list.append(divisions_list[team.division])
        elif team.conference == "East":
            teams_east_list.append(team)
            if not divisions_list[team.division] in divisions_east_list:
                divisions_east_list.append(divisions_list[team.division])
    west_conference = Conference("West", teams_west_list, divisions_west_list)
    east_conference = Conference("East", teams_east_list, divisions_east_list)
    conference_list["West"] = west_conference
    conference_list["East"] = east_conference
    return conference_list

def generateGameData(games):
    for row in games.iter_rows(range_string="A2:E1231"):
        game = (row[1].value, row[2].value, [row[3].value, row[4].value])
        GAME_DATA[row[0].value].append(game)

def canQualify(at_risk_team, teams_dict, games_remaining):
    if len(games_remaining) == 0:
        teams_list = []
        for team_name in teams_dict.keys():
            teams_list.append(teams_dict[team_name])
        divisions_dict = generateDivisions(teams_list)
        division_list = []
        for division_name in divisions_dict.keys():
            division_list.append(divisions_dict[division_name])
        conference_dict = generateConference(teams_list, division_list)
        conference = conference_dict[at_risk_team.conference]
        conference.rankTeams()
        for i in range(1, 8):
            if conference.rankings[i].name == at_risk_team.name:
                return True
        return False
    else:
        (team1, team2, outcome) = games_remaining.pop(0)
        teams_dict_win = copy.deepcopy(teams_dict)
        teams_dict_lose = copy.deepcopy(teams_dict)
        teams_dict_win[team1].addGame(teams_dict_win[team2], (1, 0))
        teams_dict_win[team2].addGame(teams_dict_win[team1], (0, 1))
        teams_dict_lose[team1].addGame(teams_dict_win[team2], (0, 1))
        teams_dict_lose[team2].addGame(teams_dict_win[team1], (1, 0))
        return canQualify(at_risk_team, teams_dict_win, games_remaining) or canQualify(at_risk_team, teams_dict_lose, games_remaining)
        
        

def checkElimination(at_risk_team, current_date, conference):
    """
    copy all teams
    for all games with team being checked, assume won
    for all games with 8th seed, assume lost
    store rest of games in array [(team1, team2, outcome)]
    systematically iterate through rest of game outcomes
        for every set of game outcomes create copy of division and conference and rankTeams
        if team being checked can rank higher than 8th seed return playoffs
        else return date
    """
    at_risk_team = copy.deepcopy(at_risk_team)
    teams_copy = {}
    other_games = []
    for team in WESTERN_CONF.teams:
        teams_copy[team.name] = copy.deepcopy(team)
    for team in EASTERN_CONF.teams:
        teams_copy[team.name] = copy.deepcopy(team)
    for date in GAME_DATA.keys():
        for game in GAME_DATA[date]:
            if at_risk_team.name in game:
                opponent = game[0]
                if opponent == at_risk_team.name:
                    opponent = game[1]
                opponent = teams_copy[opponent]
                at_risk_team.addGame(opponent, (100, 0))
                opponent.addGame(at_risk_team, (0, 100))
            else:
                other_games.append(game)
    if not canQualify(at_risk_team, teams_copy, other_games):
        return date
    return "Playoffs"
    
        
    """ end of edit """    
    
#     if at_risk_team.conference_rank > 8:
#         eighth_seed = conference.rankings[8].loseRest()
#         team_possible = at_risk_team.winRest()
#         if conference == WESTERN_CONF:
#             rank = conference.settleTie([eighth_seed, team_possible], conference, EASTERN_CONF)
#         else:
#             rank = conference.settleTie([eighth_seed, team_possible], conference, WESTERN_CONF)
#         if rank[0] == eighth_seed:
#             return current_date.strftime("%m/%d/%Y")
#     return "Playoffs"

def updateSeason(date, (first_team, second_team, score), teams, divisions, conferences, at_least_41_games_played):
    team1 = teams[first_team]
    team2 = teams[second_team]
    team1.addGame(team2, score)
    team2.addGame(team1, [score[1], score[0]])
    if not at_least_41_games_played and (team1.games_played == 41 or team2.games_played == 41):
        at_least_41_games_played = True
    if at_least_41_games_played:
        for conference_names in conferences:
            conferences[conference_names].rankTeams()
    return at_least_41_games_played

def writeToSheet(ws, teams):
    for row in ws.iter_rows(range_string="A2:B31"):
        row[1].value = teams[row[0].value].eliminated

def main():
    global WESTERN_CONF
    global EASTERN_CONF
    print "Analyzing data..."
    wb = load_workbook('Analytics_Attachment.xlsx')
    ws = wb['Division_Info']
    games = wb['2016_17_NBA_Scores']
    teams = generateTeams(ws)
    divisions = generateDivisions(teams)
    conferences = generateConference(teams, divisions)
    WESTERN_CONF = conferences["West"]
    EASTERN_CONF = conferences["East"]
    generateGameData(games)
    at_least_41_games_played = False

    for date in sorted(GAME_DATA.keys()):
        for game in GAME_DATA.pop(date, None):
            at_least_41_games_played = updateSeason(date, game, teams, divisions, conferences, at_least_41_games_played)
        for team in WESTERN_CONF.teams:
            if team.eliminated == "Playoffs":
                team.eliminated = checkElimination(team, date, WESTERN_CONF)
        for team in EASTERN_CONF.teams:
            if team.eliminated == "Playoffs":
                team.eliminated = checkElimination(team, date, EASTERN_CONF)
                
    #writing to Analytics_Attachment.xlsx tab "NBA_Clinch_Dates"
    ws = wb['NBA_Clinch_Dates']
    writeToSheet(ws, teams)
    wb.save('Analytics_Attachment.xlsx')
    
    print "Finished."

if __name__ == "__main__":
    main()
