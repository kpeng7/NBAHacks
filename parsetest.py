from openpyxl import load_workbook
from collections import defaultdict
from p2 import *

wb = load_workbook('Analytics_Attachment.xlsx')

print wb.get_sheet_names()
opponents = []
ws = wb['Division_Info']
games = wb['2016_17_NBA_Scores']
team_list = {}
division_list = {}
conference_list = {}
GAME_DATA = defaultdict(list)

for row in ws.iter_rows(range_string="A2:C31"):
    opponents.append(row[0].value)
    
#generateTeams
for row in ws.iter_rows(range_string="A2:C31"):
    team1 = Team(row[0].value, row[1].value, row[2].value, opponents)
    team_list[row[0].value] = team1
print team_list.keys()

#generateDivisions
for team_name in team_list:
    team = team_list[team_name]
    current_division = team.division
    if not division_list.has_key(current_division):
        teams_division_list = []
        for team_name_temp in team_list:
            team_temp = team_list[team_name_temp]
            if team_temp.division == current_division:
                teams_division_list.append(team_temp)
        division_temp = Division(current_division, teams_division_list)
        division_list[current_division] = division_temp
print division_list.keys()

#generateConference
teams_west_list = []
teams_east_list = []
for team_name in team_list:
    team = team_list[team_name]
    if team.conference == "West":
        teams_west_list.append(team)
    elif team.conference == "East":
        teams_east_list.append(team)
west_conference = Conference("West", teams_west_list)
east_conference = Conference("East", teams_east_list)
conference_list["West"] = west_conference
conference_list["East"] = east_conference
print conference_list.keys()

#generateGame_Data
for row in games.iter_rows(range_string="A2:E1231"):
    game = (row[1].value, row[2].value, (row[3].value, row[4].value))
    GAME_DATA[row[0].value].append(game)
print GAME_DATA
print GAME_DATA.keys()

